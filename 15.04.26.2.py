# importamos load_workbook
from openpyxl import load_workbook
# ruta de nuestro archivo
filesheet = "./demosheet.xlsx"
# creamos el objeto load_workbook
wb = load_workbook(filesheet)
# seleccionamos el archivo
sheet = wb.active
# escribimos los datos con sus respectivas filas y columnas
datos = [('id','nombre','edad'),
         (0,"Jose",35),
          (1,"Carlos",27),
           (2,"Sofía",24)]
# recorremos las columnas y
for row in datos:
    sheet.append(row)
#guardamos los cambios
wb.save(filesheet)
