# importamos load_workbook
from openpyxl import load_workbook
# ruta de nuestro archivo
filesheet = "./demosheet.xlsx"
# creamos el objeto load_workbook
wb = load_workbook(filesheet)
# seleccionamos el archivo
sheet = wb.active
# obtenemos el valor de la celda A1
A1 = sheet['A1'].value
# obtenemos el valor de la celda B3
B3 = sheet['B5'].value
# obtenemos el valor de la celda C5
C5 = sheet['C5'].value
# mostramos los valores
celdas = [A1, B3, C5]
for valor in celdas:
    print(valor)